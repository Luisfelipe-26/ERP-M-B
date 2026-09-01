from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, cast, Date
from database import get_db
import models, schemas, auth
from models import ESTADOS_OT
from routers.sequences import get_next, peek_next
from routers.contabilidad import _crear_asiento_auto, _get_regla_cuentas, _actualizar_saldos
from typing import List, Optional
from datetime import datetime
from decimal import Decimal
import audit

router = APIRouter(prefix="/api/ordenes", tags=["ordenes"])


def _next_ot_id(db: Session) -> int:
    """Use atomic sequence for OT numbering."""
    return int(get_next("OT", db))


def _recalc_orden(orden: models.OrdenTrabajo, db: Session):
    """Recalculates and persists costo totals for an OT (MO + Insumos + Equipo)."""
    mo_total = db.query(func.coalesce(func.sum(models.OTManoObra.costo_mo), 0)).filter(
        models.OTManoObra.ot_id == orden.ot_id
    ).scalar() or 0
    ins_total = db.query(func.coalesce(func.sum(models.OTDetalle.costo_real), 0)).filter(
        models.OTDetalle.ot_id == orden.ot_id
    ).scalar() or 0
    campo = db.query(models.Campo).filter(models.Campo.id_campo == orden.campo_id).first()
    area = campo.area_ha if campo and campo.area_ha is not None and campo.area_ha > 0 else 1
    # Costo equipo
    equipo_total = round(float(orden.horas_equipo or 0) * float(orden.tarifa_equipo or 0), 2)
    orden.costo_mano_obra = round(float(mo_total), 2)
    orden.costo_insumos = round(float(ins_total), 2)
    orden.costo_equipo = equipo_total
    orden.costo_total = round(float(mo_total) + float(ins_total) + equipo_total, 2)
    orden.costo_ha = round(orden.costo_total / area, 2)


def _crear_movimiento_consumo_ot(db: Session, ot_id: int, producto: models.Producto,
                                  cantidad: float, costo_unitario: float, fecha, usuario_id: int):
    """
    H-1 FIX: Genera un MovimientoInventario tipo OT por cada línea de insumo consumido.
    Esto garantiza trazabilidad completa en el kardex.
    """
    nuevo_stock = max(0.0, (producto.stock_actual or 0) - cantidad)
    # Inventario manda: el consumo sale valuado al costo promedio ponderado del
    # producto (Valor INV). La OTDetalle se alinea a este mismo valor.
    cp = producto.costo_promedio or producto.costo_unitario or 0

    mov = models.MovimientoInventario(
        num_documento=f"OT-{ot_id}",
        producto_id=producto.id_prod,
        tipo_doc="OT",
        tipo="salida",
        motivo="Consumo OT",
        cantidad=cantidad,
        costo_unitario=round(cp, 4),
        costo_promedio_post=round(cp, 4),
        stock_post=nuevo_stock,
        ot_referencia=ot_id,
        observacion=f"Consumo automático - OT #{ot_id}",
        fecha=fecha or datetime.now(),
        usuario_id=usuario_id,
    )
    db.add(mov)
    producto.stock_actual = nuevo_stock
    return mov


def _crear_movimiento_reversion_ot(db: Session, ot_id: int, producto: models.Producto,
                                    cantidad: float, usuario_id: int, costo_unitario: float = None):
    """
    H-2 FIX: Genera un MovimientoInventario de reversión al eliminar una OT.
    Restaura stock con documento auditable.
    Valúa al mismo costo que la salida original (costo_unitario) para netear a cero.
    """
    nuevo_stock = (producto.stock_actual or 0) + cantidad
    cp = producto.costo_promedio or producto.costo_unitario or 0
    costo_mov = costo_unitario if costo_unitario else cp

    mov = models.MovimientoInventario(
        num_documento=f"OT-{ot_id}-REV",
        producto_id=producto.id_prod,
        tipo_doc="OT",
        tipo="entrada",
        motivo="Reversión OT eliminada",
        cantidad=cantidad,
        costo_unitario=round(costo_mov, 4),
        costo_promedio_post=round(cp, 4),
        stock_post=nuevo_stock,
        ot_referencia=ot_id,
        observacion=f"Reversión por eliminación de OT #{ot_id}",
        fecha=datetime.now(),
        usuario_id=usuario_id,
    )
    db.add(mov)
    producto.stock_actual = nuevo_stock
    return mov


def _lineas_cierre_ot(db: Session, orden: models.OrdenTrabajo) -> list:
    """Construye las líneas del asiento de cierre de una OT desde sus costos actuales.
    Única fuente para el cierre y la re-valuación, así ambos generan lo mismo."""
    ot_id = orden.ot_id
    lineas = []
    costo_mo = Decimal(str(orden.costo_mano_obra or 0))
    costo_ins = Decimal(str(orden.costo_insumos or 0))
    costo_eq = Decimal(str(orden.costo_equipo or 0))
    un_id = getattr(orden, "unidad_negocio_id", None)
    dep_id = getattr(orden, "departamento_id", None)

    r_mo = _get_regla_cuentas(db, "nomina", "salario_jornada")
    if r_mo and costo_mo > 0:
        lineas.append({"cuenta_id": r_mo[0], "debe": costo_mo, "haber": 0,
                       "campo_id": orden.campo_id, "unidad_negocio_id": un_id, "departamento_id": dep_id,
                       "descripcion_linea": f"MO directa OT #{ot_id}"})
        lineas.append({"cuenta_id": r_mo[1], "debe": 0, "haber": costo_mo,
                       "descripcion_linea": f"Nóminas por pagar OT #{ot_id}"})

    r_ins = _get_regla_cuentas(db, "consumo_ot", "salida_insumo")
    if r_ins and costo_ins > 0:
        lineas.append({"cuenta_id": r_ins[0], "debe": costo_ins, "haber": 0,
                       "campo_id": orden.campo_id, "unidad_negocio_id": un_id, "departamento_id": dep_id,
                       "descripcion_linea": f"Insumos OT #{ot_id}"})
        lineas.append({"cuenta_id": r_ins[1], "debe": 0, "haber": costo_ins,
                       "descripcion_linea": f"Salida inventario OT #{ot_id}"})

    if costo_eq > 0:
        cta_eq = db.query(models.CuentaContable).filter(
            models.CuentaContable.codigo == "5.2.02").first()
        cta_cxp = db.query(models.CuentaContable).filter(
            models.CuentaContable.codigo == "2.1.01.01").first()
        if cta_eq and cta_cxp:
            lineas.append({"cuenta_id": cta_eq.id, "debe": costo_eq, "haber": 0,
                           "campo_id": orden.campo_id, "unidad_negocio_id": un_id, "departamento_id": dep_id,
                           "descripcion_linea": f"Equipo OT #{ot_id}"})
            lineas.append({"cuenta_id": cta_cxp.id, "debe": 0, "haber": costo_eq,
                           "descripcion_linea": f"CxP equipo OT #{ot_id}"})
    return lineas


def _revaluar_asiento_ot(db: Session, orden: models.OrdenTrabajo):
    """Re-valúa el asiento de cierre de una OT a los costos actuales de la orden:
    reconstruye las líneas con _lineas_cierre_ot y ajusta saldos mensuales en ambas
    direcciones (los asientos automáticos son gestionados por el sistema).
    Devuelve None si no hay asiento vigente; {"duplicados": [...]} si hay más de uno
    (no se toca ninguno); {"asiento": num, "anulado": bool} si se actualizó."""
    asientos = db.query(models.AsientoContable).filter(
        models.AsientoContable.origen == "OT",
        models.AsientoContable.referencia_id == str(orden.ot_id),
        models.AsientoContable.estado != "anulado",
    ).order_by(models.AsientoContable.id.asc()).all()
    if not asientos:
        return None
    if len(asientos) > 1:
        return {"duplicados": [a.numero for a in asientos]}
    asiento = asientos[0]

    lineas_data = [l for l in _lineas_cierre_ot(db, orden)
                   if Decimal(str(l.get("debe") or 0)) > 0 or Decimal(str(l.get("haber") or 0)) > 0]

    _actualizar_saldos(db, asiento, -1)
    db.query(models.LineaAsiento).filter(
        models.LineaAsiento.asiento_id == asiento.id
    ).delete(synchronize_session="fetch")

    if not lineas_data:
        asiento.total_debe = 0
        asiento.total_haber = 0
        asiento.estado = "anulado"
        return {"asiento": asiento.numero, "anulado": True}

    for l in lineas_data:
        db.add(models.LineaAsiento(
            asiento_id=asiento.id, cuenta_id=l["cuenta_id"],
            debe=Decimal(str(l.get("debe") or 0)),
            haber=Decimal(str(l.get("haber") or 0)),
            campo_id=l.get("campo_id"), tercero_id=l.get("tercero_id"),
            descripcion_linea=l.get("descripcion_linea"),
            unidad_negocio_id=l.get("unidad_negocio_id"),
            departamento_id=l.get("departamento_id"),
            almacen_id=l.get("almacen_id"),
        ))
    asiento.total_debe = sum(Decimal(str(l.get("debe") or 0)) for l in lineas_data)
    asiento.total_haber = sum(Decimal(str(l.get("haber") or 0)) for l in lineas_data)
    db.flush()
    db.expire(asiento, ["lineas"])
    _actualizar_saldos(db, asiento, 1)
    return {"asiento": asiento.numero, "anulado": False}


# ─── Preview next OT ID (must be BEFORE /{ot_id} to avoid route conflict) ────
@router.get("/preview/next-id")
def next_ot_id_preview(db: Session = Depends(get_db), _=Depends(auth.get_current_user)):
    """Preview the next OT number without consuming it."""
    return {"next_ot_id": int(peek_next("OT", db))}


@router.put("/sequence/reset")
def reset_ot_sequence(nuevo_inicio: int = Query(0, description="Último número usado (la próxima OT será este +1)"),
                      db: Session = Depends(get_db),
                      current_user=Depends(auth.require_admin)):
    seq = db.query(models.Sequence).filter(models.Sequence.tipo == "OT").first()
    if not seq:
        seq = models.Sequence(tipo="OT", ultimo_numero=nuevo_inicio)
        db.add(seq)
    else:
        seq.ultimo_numero = nuevo_inicio
    db.commit()
    next_id = int(peek_next("OT", db))
    audit.log(db, current_user, "CONFIG", "SEQUENCE", "OT",
              f"Secuencia OT reseteada a {nuevo_inicio}, próxima OT será #{next_id}")
    return {"message": f"Secuencia OT reseteada. Próxima OT será #{next_id}", "next_ot_id": next_id}


@router.post("/recalcular-costos")
def recalcular_costos_ot(db: Session = Depends(get_db),
                          current_user=Depends(auth.require_admin)):
    """Recalculate costo_mano_obra for all OTs from actual OTManoObra records."""
    ordenes = db.query(models.OrdenTrabajo).all()
    fixed = 0
    for orden in ordenes:
        old_mo = float(orden.costo_mano_obra or 0)
        _recalc_orden(orden, db)
        if round(old_mo, 2) != round(float(orden.costo_mano_obra or 0), 2):
            fixed += 1
    db.commit()
    return {"message": f"Recalculadas {len(ordenes)} OTs, {fixed} tenían costos desactualizados"}


@router.get("", response_model=List[schemas.OrdenTrabajoOut])
def list_ordenes(
    estado: Optional[str] = None,
    campo_id: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    buscar: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_db),
    _=Depends(auth.get_current_user)
):
    q = db.query(models.OrdenTrabajo)
    if estado:
        q = q.filter(models.OrdenTrabajo.estado == estado)
    if campo_id:
        q = q.filter(models.OrdenTrabajo.campo_id == campo_id)
    if mes and ano:
        q = q.filter(
            models.OrdenTrabajo.fecha_ejecucion.isnot(None),
            extract('month', models.OrdenTrabajo.fecha_ejecucion) == mes,
            extract('year', models.OrdenTrabajo.fecha_ejecucion) == ano
        )
    if buscar:
        q = q.filter(models.OrdenTrabajo.supervisor.ilike(f"%{buscar}%"))
    ordenes = q.order_by(models.OrdenTrabajo.ot_id.asc()).offset(skip).limit(limit).all()
    act_ids = {o.actividad_id for o in ordenes if o.actividad_id}
    act_map = {}
    if act_ids:
        filas = db.query(models.Actividad.id_act, models.Actividad.actividad).filter(
            models.Actividad.id_act.in_(act_ids)).all()
        act_map = {r.id_act: r.actividad for r in filas}
    out = []
    for o in ordenes:
        d = schemas.OrdenTrabajoOut.from_orm(o)
        d.actividad_nombre = act_map.get(o.actividad_id)
        out.append(d)
    return out


@router.post("", response_model=schemas.OrdenTrabajoOut)
def create_orden(data: schemas.OrdenTrabajoCreate, db: Session = Depends(get_db),
                  current_user: models.Usuario = Depends(auth.get_current_user)):
    # Validate foreign keys
    campo = None
    if data.campo_id:
        campo = db.query(models.Campo).filter(models.Campo.id_campo == data.campo_id, models.Campo.activo == True).first()
        if not campo:
            raise HTTPException(status_code=400, detail=f"Campo '{data.campo_id}' no existe o está inactivo")

    actividad = db.query(models.Actividad).filter(models.Actividad.id_act == data.actividad_id, models.Actividad.activo == True).first()
    if not actividad:
        raise HTTPException(status_code=400, detail=f"Actividad '{data.actividad_id}' no existe o está inactiva")

    if data.ot_id and db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == data.ot_id).first():
        raise HTTPException(status_code=400, detail=f"Ya existe una OT con ID {data.ot_id}")

    ot_id = data.ot_id or _next_ot_id(db)

    orden = models.OrdenTrabajo(
        ot_id=ot_id,
        fecha_ejecucion=data.fecha_ejecucion,
        hora_inicio=data.hora_inicio,
        campo_id=data.campo_id or None,  # empty string → NULL for FK
        actividad_id=data.actividad_id,
        supervisor=data.supervisor,
        estado=data.estado,
        equipo=data.equipo,
        horas_equipo=data.horas_equipo,
        tarifa_equipo=data.tarifa_equipo,
        observaciones=data.observaciones,
    )
    db.add(orden)
    db.flush()

    # ── Mano de Obra ──
    costo_mo_total = 0.0
    warnings = []
    fecha_ot = data.fecha_ejecucion.strftime('%Y-%m-%d') if data.fecha_ejecucion else None
    for mo_data in (data.mano_obra or []):
        if not mo_data.trabajador_id:
            continue
        trabajador = db.query(models.Trabajador).filter(
            models.Trabajador.id_trab == mo_data.trabajador_id, models.Trabajador.activo == True
        ).first()
        if not trabajador:
            raise HTTPException(status_code=400, detail=f"Trabajador '{mo_data.trabajador_id}' no existe")

        # H-17 FIX: Alertar si el trabajador ya tiene jornada en otra OT en la misma fecha/hora
        if fecha_ot:
            existing = db.query(models.OTManoObra).join(models.OrdenTrabajo).filter(
                models.OTManoObra.trabajador_id == mo_data.trabajador_id,
                cast(models.OrdenTrabajo.fecha_ejecucion, Date) == fecha_ot,
                models.OrdenTrabajo.ot_id != ot_id,
            ).first()
            if existing:
                warnings.append(
                    f"⚠ '{trabajador.nombre}' ya tiene jornada en OT #{existing.ot_id} "
                    f"para la fecha {fecha_ot}. Verifique para evitar pago doble."
                )

        costo_hora = mo_data.costo_hora or trabajador.costo_hora or 62.50
        horas = mo_data.horas_netas or 8.0
        if float(horas) > 16.0 and (mo_data.modalidad or 'Jornada') != 'Ajuste':
            raise HTTPException(status_code=400,
                detail=f"Trabajador '{trabajador.nombre}': horas ({horas}) exceden el máximo de 16. Verifique el dato.")
        horas = max(0.0, float(horas))
        costo_mo = mo_data.costo_mo if mo_data.costo_mo is not None else (costo_hora * horas)
        costo_mo = max(0.0, float(costo_mo))
        mo = models.OTManoObra(
            ot_id=ot_id,
            trabajador_id=mo_data.trabajador_id,
            fecha=mo_data.fecha or data.fecha_ejecucion,
            hora_inicio=mo_data.hora_inicio,
            hora_fin=mo_data.hora_fin,
            pausa_min=mo_data.pausa_min,
            horas_netas=horas,
            modalidad=mo_data.modalidad,
            costo_hora=costo_hora,
            costo_mo=costo_mo,
        )
        db.add(mo)
        costo_mo_total += costo_mo

    # ── Insumos ──
    costo_insumos_total = 0.0
    for det_data in (data.detalles or []):
        if not det_data.producto_id:
            continue
        prod = db.query(models.Producto).filter(
            models.Producto.id_prod == det_data.producto_id, models.Producto.activo == True
        ).first()
        if not prod:
            raise HTTPException(status_code=400, detail=f"Producto '{det_data.producto_id}' no existe")

        cantidad = max(0.0, float(det_data.cantidad_usada or 0))

        # H-6 FIX: Validar stock suficiente antes de consumir (solo inventariables)
        if cantidad > 0 and prod.es_inventariable and (prod.stock_actual or 0) < cantidad:
            raise HTTPException(status_code=400,
                detail=f"Stock insuficiente para '{prod.producto}'. "
                       f"Disponible: {prod.stock_actual} {prod.unidad}, Requerido: {cantidad} {prod.unidad}")

        # Inventario manda (Valor INV): los consumos inventariables se valúan al costo
        # promedio del inventario, no al costo tecleado. Los no inventariables no pasan
        # por inventario, así que usan el costo tecleado o el de lista.
        if prod.es_inventariable:
            costo_unit = prod.costo_promedio or prod.costo_unitario or 0
        else:
            costo_unit = det_data.costo_unitario if det_data.costo_unitario is not None else (prod.costo_unitario or 0)
        costo_real = max(0.0, round(cantidad * costo_unit, 2))

        det = models.OTDetalle(
            ot_id=ot_id,
            producto_id=det_data.producto_id,
            fecha=det_data.fecha or data.fecha_ejecucion,
            cantidad_usada=cantidad,
            cantidad_tanque=det_data.cantidad_tanque,
            unidad=det_data.unidad or prod.unidad or "unidad",
            costo_unitario=costo_unit,
            costo_real=costo_real,
            observacion=det_data.observacion,
        )
        db.add(det)
        costo_insumos_total += costo_real

        # H-1 FIX: Generar movimiento de inventario por cada consumo de OT (solo inventariables)
        if cantidad > 0 and prod.es_inventariable:
            _crear_movimiento_consumo_ot(
                db, ot_id, prod, cantidad, costo_unit,
                det_data.fecha or data.fecha_ejecucion,
                current_user.id
            )

    # FIX: campo puede ser None si campo_id es vacío (OTs generales/vivero)
    area = (campo.area_ha if campo and campo.area_ha is not None and campo.area_ha > 0 else 1)
    campo_nombre = campo.nombre if campo else "General"
    # H-13 FIX: Redondeo consistente a 2 decimales
    costo_mo_total = round(costo_mo_total, 2)
    costo_insumos_total = round(costo_insumos_total, 2)
    costo_equipo = round(float(data.horas_equipo or 0) * float(data.tarifa_equipo or 0), 2)
    costo_total = round(costo_mo_total + costo_insumos_total + costo_equipo, 2)
    orden.costo_mano_obra = costo_mo_total
    orden.costo_insumos = costo_insumos_total
    orden.costo_equipo = costo_equipo
    orden.costo_total = costo_total
    orden.costo_ha = round(costo_total / area, 2)
    orden.horas_mo = round(sum(max(0, float(m.horas_netas or 8)) for m in (data.mano_obra or []) if m.trabajador_id), 2)

    audit.log(db, current_user, "CREAR", "OT", str(ot_id),
              f"OT #{ot_id} creada: {campo_nombre} / {actividad.actividad} — Total: RD$ {costo_total:,.2f}",
              {"campo": data.campo_id, "actividad": data.actividad_id, "costo_total": costo_total,
               "costo_mo": costo_mo_total, "costo_insumos": costo_insumos_total, "costo_equipo": costo_equipo})

    try:
        db.commit()
        db.refresh(orden)
    except Exception:
        db.rollback()
        import logging
        logging.getLogger(__name__).exception("Error creando OT %s", ot_id)
        raise HTTPException(500, "Error al crear la orden de trabajo")
    result = schemas.OrdenTrabajoOut.model_validate(orden).model_dump()
    if warnings:
        result["warnings"] = warnings
    return result


@router.put("/{ot_id}")
def update_orden(ot_id: int, data: schemas.OrdenTrabajoCreate, db: Session = Depends(get_db),
                  current_user: models.Usuario = Depends(auth.get_current_user)):
    """Update an existing OT — replaces MO and detalles, recalculates costs."""
    orden = db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == ot_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if orden.estado == "Cerrada" and current_user.rol not in ("admin", "supervisor", "operador"):
        raise HTTPException(status_code=403, detail="No tiene permisos para editar una orden cerrada")

    campo = None
    if data.campo_id:
        campo = db.query(models.Campo).filter(models.Campo.id_campo == data.campo_id, models.Campo.activo == True).first()
        if not campo:
            raise HTTPException(status_code=400, detail=f"Campo '{data.campo_id}' no existe")
    actividad = db.query(models.Actividad).filter(models.Actividad.id_act == data.actividad_id, models.Actividad.activo == True).first()
    if not actividad:
        raise HTTPException(status_code=400, detail=f"Actividad '{data.actividad_id}' no existe")

    # Restore stock from old detalles (no reversion movement — edit, not delete)
    # Solo inventariables: a los servicios nunca se les descontó stock, restaurarlo
    # aquí inflaba stock fantasma en cada edición.
    old_detalles = db.query(models.OTDetalle).filter(models.OTDetalle.ot_id == ot_id).all()
    for det in old_detalles:
        if det.producto_id and det.cantidad_usada and det.cantidad_usada > 0:
            prod = db.query(models.Producto).filter(models.Producto.id_prod == det.producto_id).first()
            if prod and prod.es_inventariable:
                prod.stock_actual = (prod.stock_actual or 0) + det.cantidad_usada

    # Delete old MO and detalles
    db.query(models.OTManoObra).filter(models.OTManoObra.ot_id == ot_id).delete()
    db.query(models.OTDetalle).filter(models.OTDetalle.ot_id == ot_id).delete()
    # Delete ALL old OT inventory movements (consumos + any orphaned reversions)
    db.query(models.MovimientoInventario).filter(
        models.MovimientoInventario.num_documento.in_([f"OT-{ot_id}", f"OT-{ot_id}-REV"]),
        models.MovimientoInventario.tipo_doc == "OT"
    ).delete(synchronize_session="fetch")

    # Update header fields
    orden.fecha_ejecucion = data.fecha_ejecucion
    orden.hora_inicio = data.hora_inicio
    orden.campo_id = data.campo_id or None  # empty string → NULL for FK
    orden.actividad_id = data.actividad_id
    orden.supervisor = data.supervisor
    orden.estado = data.estado
    orden.equipo = data.equipo
    orden.horas_equipo = data.horas_equipo
    orden.tarifa_equipo = data.tarifa_equipo
    orden.observaciones = data.observaciones

    # Re-add MO
    costo_mo_total = 0.0
    for mo_data in (data.mano_obra or []):
        if not mo_data.trabajador_id:
            continue
        costo_mo = max(0.0, float(mo_data.costo_mo or 0))
        mo = models.OTManoObra(
            ot_id=ot_id,
            trabajador_id=mo_data.trabajador_id,
            fecha=mo_data.fecha or data.fecha_ejecucion,
            hora_inicio=mo_data.hora_inicio,
            hora_fin=mo_data.hora_fin,
            pausa_min=mo_data.pausa_min,
            horas_netas=float(mo_data.horas_netas or 8),
            modalidad=mo_data.modalidad,
            costo_hora=float(mo_data.costo_hora or 0),
            costo_mo=costo_mo,
        )
        db.add(mo)
        costo_mo_total += costo_mo

    # Re-add detalles with new inventory movements
    costo_insumos_total = 0.0
    for det_data in (data.detalles or []):
        if not det_data.producto_id:
            continue
        prod = db.query(models.Producto).filter(
            models.Producto.id_prod == det_data.producto_id, models.Producto.activo == True
        ).first()
        if not prod:
            raise HTTPException(status_code=400, detail=f"Producto '{det_data.producto_id}' no existe")
        cantidad = max(0.0, float(det_data.cantidad_usada or 0))
        if cantidad > 0 and prod.es_inventariable and (prod.stock_actual or 0) < cantidad:
            raise HTTPException(status_code=400,
                detail=f"Stock insuficiente para '{prod.producto}'. Disponible: {prod.stock_actual}")
        # Inventario manda (Valor INV): inventariables al costo promedio; no
        # inventariables al costo tecleado o de lista.
        if prod.es_inventariable:
            costo_unit = prod.costo_promedio or prod.costo_unitario or 0
        else:
            costo_unit = det_data.costo_unitario if det_data.costo_unitario is not None else (prod.costo_unitario or 0)
        costo_real = max(0.0, round(cantidad * costo_unit, 2))

        det = models.OTDetalle(
            ot_id=ot_id, producto_id=det_data.producto_id,
            fecha=det_data.fecha or data.fecha_ejecucion,
            cantidad_usada=cantidad, unidad=det_data.unidad or prod.unidad,
            costo_unitario=costo_unit, costo_real=costo_real,
        )
        db.add(det)
        costo_insumos_total += costo_real
        if cantidad > 0 and prod.es_inventariable:
            _crear_movimiento_consumo_ot(db, ot_id, prod, cantidad, costo_unit,
                det_data.fecha or data.fecha_ejecucion, current_user.id)

    # Recalculate totals
    area = (campo.area_ha if campo and campo.area_ha and campo.area_ha > 0 else 1)
    costo_equipo = round(float(data.horas_equipo or 0) * float(data.tarifa_equipo or 0), 2)
    orden.costo_mano_obra = round(costo_mo_total, 2)
    orden.costo_insumos = round(costo_insumos_total, 2)
    orden.costo_equipo = costo_equipo
    orden.costo_total = round(costo_mo_total + costo_insumos_total + costo_equipo, 2)
    orden.costo_ha = round(orden.costo_total / area, 2)

    # Si la OT ya tiene asiento de cierre, re-valuarlo a los nuevos costos para
    # que OT ↔ Asientos no diverja al editar una orden cerrada.
    res_asiento = _revaluar_asiento_ot(db, orden)

    audit.log(db, current_user, "EDITAR", "OT", str(ot_id),
              f"OT #{ot_id} editada — Total: RD$ {orden.costo_total:,.2f}",
              {"asiento_revaluado": res_asiento} if res_asiento else None)

    try:
        db.commit()
    except Exception:
        db.rollback()
        import logging
        logging.getLogger(__name__).exception("Error editando OT %s", ot_id)
        raise HTTPException(500, "Error al editar la orden de trabajo")
    db.refresh(orden)
    return orden


@router.get("/{ot_id}", response_model=schemas.OrdenTrabajoOut)
def get_orden(ot_id: int, db: Session = Depends(get_db), _=Depends(auth.get_current_user)):
    orden = db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == ot_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    return orden


@router.get("/{ot_id}/detalle")
def get_orden_detalle(ot_id: int, db: Session = Depends(get_db), _=Depends(auth.get_current_user)):
    orden = db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == ot_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    mano_obra = db.query(models.OTManoObra).filter(models.OTManoObra.ot_id == ot_id).all()
    detalles = db.query(models.OTDetalle).filter(models.OTDetalle.ot_id == ot_id).all()

    asiento_info = None
    if orden.estado == "Cerrada":
        asiento = db.query(models.AsientoContable).filter(
            models.AsientoContable.origen == "OT",
            models.AsientoContable.referencia_id == str(ot_id),
        ).first()
        if asiento:
            asiento_info = {"numero": asiento.numero, "fecha": str(asiento.fecha),
                            "total_debe": float(asiento.total_debe or 0),
                            "estado": asiento.estado}

    return {
        "orden": schemas.OrdenTrabajoOut.model_validate(orden),
        "mano_obra": [schemas.OTManoObraOut.model_validate(m) for m in mano_obra],
        "detalles": [schemas.OTDetalleOut.model_validate(d) for d in detalles],
        "asiento_contable": asiento_info,
    }


@router.patch("/{ot_id}/mano-obra/{mo_id}")
def update_mano_obra(ot_id: int, mo_id: int, data: dict,
                     db: Session = Depends(get_db),
                     current_user: models.Usuario = Depends(auth.get_current_user)):
    orden = db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == ot_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if orden.estado == "Cerrada" and current_user.rol not in ("admin", "supervisor", "operador"):
        raise HTTPException(status_code=403, detail="No tiene permisos para editar una orden cerrada")
    mo = db.query(models.OTManoObra).filter(
        models.OTManoObra.id == mo_id, models.OTManoObra.ot_id == ot_id
    ).first()
    if not mo:
        raise HTTPException(status_code=404, detail="Registro de mano de obra no encontrado")
    allowed = {"hora_inicio", "hora_fin", "pausa_min", "horas_netas", "costo_hora", "costo_mo", "modalidad", "fecha"}
    for k, v in data.items():
        if k in allowed:
            setattr(mo, k, v)
    if "horas_netas" in data and "costo_hora" in data and "costo_mo" not in data:
        mo.costo_mo = round(float(data["horas_netas"]) * float(data["costo_hora"]), 2)
    _recalc_orden(orden, db)
    audit.log(db, current_user, "EDITAR_MO", "OT", str(ot_id),
              f"MO #{mo_id} editada en OT #{ot_id} — Costo: RD$ {mo.costo_mo:,.2f}")
    db.commit()
    return {"ok": True, "costo_mo": float(mo.costo_mo), "costo_total_ot": float(orden.costo_total)}


@router.put("/{ot_id}/estado")
def update_estado(ot_id: int, estado: str = Query(...), hora_cierre: Optional[str] = Query(None),
                   db: Session = Depends(get_db),
                   current_user: models.Usuario = Depends(auth.get_current_user)):
    if estado not in ESTADOS_OT:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Valores permitidos: {ESTADOS_OT}")
    orden = db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == ot_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    estado_anterior = orden.estado
    orden.estado = estado
    asiento_num = None
    try:
        if estado == "Cerrada":
            orden.fecha_cierre = datetime.now()
            orden.hora_cierre = hora_cierre or datetime.now().strftime("%H:%M")
            _recalc_orden(orden, db)

            fecha_ot = (orden.fecha_ejecucion.date() if isinstance(orden.fecha_ejecucion, datetime)
                        else orden.fecha_ejecucion) if orden.fecha_ejecucion else datetime.now().date()

            # Guard anti-duplicado: si la OT ya tiene asiento de cierre vigente
            # (re-cierre o reapertura+cierre), se re-valúa en lugar de crear otro.
            existente = db.query(models.AsientoContable).filter(
                models.AsientoContable.origen == "OT",
                models.AsientoContable.referencia_id == str(ot_id),
                models.AsientoContable.estado != "anulado",
            ).first()
            if existente:
                _revaluar_asiento_ot(db, orden)
                asiento_num = existente.numero
            else:
                lineas = _lineas_cierre_ot(db, orden)
                if lineas:
                    asiento = _crear_asiento_auto(
                        db, fecha_ot, "OT", str(ot_id),
                        f"Cierre OT #{ot_id} — {orden.actividad_id} / {orden.campo_id or 'General'}",
                        lineas, current_user.nombre
                    )
                    if asiento:
                        asiento_num = asiento.numero

        audit.log(db, current_user, "ESTADO", "OT", str(ot_id),
                  f"OT #{ot_id}: {estado_anterior} → {estado}",
                  {"estado_anterior": estado_anterior, "estado_nuevo": estado})
        db.commit()
    except Exception:
        db.rollback()
        import logging
        logging.getLogger(__name__).exception("Error cambiando estado OT %s", ot_id)
        raise HTTPException(500, "Error al cambiar estado de la orden")
    return {"ok": True, "estado": estado, "hora_cierre": orden.hora_cierre,
            "asiento": asiento_num}


@router.delete("/{ot_id}")
def delete_orden(ot_id: int, db: Session = Depends(get_db),
                  current_user: models.Usuario = Depends(auth.require_admin)):
    orden = db.query(models.OrdenTrabajo).filter(models.OrdenTrabajo.ot_id == ot_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    # Solo limpiar reversiones huérfanas previas (no la salida original: se conserva
    # para que quede el par salida(−) + reversión(+) = neto cero, con historial completo).
    db.query(models.MovimientoInventario).filter(
        models.MovimientoInventario.num_documento == f"OT-{ot_id}-REV",
        models.MovimientoInventario.tipo_doc == "OT"
    ).delete(synchronize_session="fetch")

    # Restaurar stock CON movimiento de reversión auditable (la salida original se mantiene).
    # Solo inventariables: los servicios nunca generaron salida — una reversión aquí
    # nacía huérfana e inflaba stock fantasma.
    detalles = db.query(models.OTDetalle).filter(models.OTDetalle.ot_id == ot_id).all()
    for det in detalles:
        if det.producto_id and det.cantidad_usada and det.cantidad_usada > 0:
            prod = db.query(models.Producto).filter(models.Producto.id_prod == det.producto_id).first()
            if prod and prod.es_inventariable:
                _crear_movimiento_reversion_ot(db, ot_id, prod, det.cantidad_usada, current_user.id,
                                               costo_unitario=det.costo_unitario)

    db.query(models.OTManoObra).filter(models.OTManoObra.ot_id == ot_id).delete()
    db.query(models.OTDetalle).filter(models.OTDetalle.ot_id == ot_id).delete()

    # Anular el asiento de cierre: el mayor no debe conservar costos de una OT borrada.
    # Se revierte su efecto en saldos mensuales y queda excluido de consultas (estado).
    asientos_ot = db.query(models.AsientoContable).filter(
        models.AsientoContable.origen == "OT",
        models.AsientoContable.referencia_id == str(ot_id),
        models.AsientoContable.estado != "anulado",
    ).all()
    for a in asientos_ot:
        _actualizar_saldos(db, a, -1)
        a.estado = "anulado"
        a.anulado_por = current_user.nombre

    audit.log(db, current_user, "ELIMINAR", "OT", str(ot_id),
              f"OT #{ot_id} eliminada: {orden.campo_id} / {orden.actividad_id} — Total era: RD$ {orden.costo_total or 0:,.2f}",
              {"campo": orden.campo_id, "actividad": orden.actividad_id, "costo_total": orden.costo_total,
               "asientos_anulados": [a.numero for a in asientos_ot]})

    db.delete(orden)
    try:
        db.commit()
    except Exception:
        db.rollback()
        import logging
        logging.getLogger(__name__).exception("Error eliminando OT %s", ot_id)
        raise HTTPException(500, "Error al eliminar la orden de trabajo")


# ── Plantilla y Carga Masiva ──

@router.get("/plantilla/descargar")
def descargar_plantilla():
    import os
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "plantilla_carga_ots.xlsx")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    return FileResponse(path, filename="plantilla_carga_ots.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/carga-masiva")
def carga_masiva(file: UploadFile = File(...), db: Session = Depends(get_db),
                 current_user: models.Usuario = Depends(auth.get_current_user)):
    import openpyxl, io
    from collections import OrderedDict
    from datetime import timedelta

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo archivos .xlsx")

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # Find sheet
    if 'OT_ManoObra' in wb.sheetnames:
        ws = wb['OT_ManoObra']
    else:
        ws = wb.active

    # Build lookup maps
    acts = {a.actividad: a.id_act for a in db.query(models.Actividad).filter(models.Actividad.activo == True).all()}
    acts_lower = {k.lower().strip(): v for k, v in acts.items()}
    trabs = {t.nombre.strip(): t.id_trab for t in db.query(models.Trabajador).filter(models.Trabajador.activo == True).all()}
    trabs_lower = {k.lower().strip(): v for k, v in trabs.items()}
    campos_set = {c.id_campo for c in db.query(models.Campo).filter(models.Campo.activo == True).all()}

    # Detect format from header row
    header_row = [str(c.value or '').strip().lower() for c in ws[1]]
    # Format A (plantilla): Fecha, Actividad, Trabajador, Hora inicio, Hora fin, Pausa, Campo, ...
    # Format B (Excel original): OT_ID, Fecha, Actividad_ID, Trabajador_ID, Hora inicio, Hora fin, Pausa, ..., Campo_ID (col 19)
    is_format_b = header_row[0] in ('ot_id', 'ot id')
    # Format A with OT grouper: first col is 'ot' (just a number, not 'ot_id')
    has_ot_col = header_row[0] in ('ot', 'ot_id', 'ot id')

    # Parse rows
    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        raw_rows.append(list(row))

    if not raw_rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos")

    def parse_fecha(v):
        if isinstance(v, datetime):
            return v.strftime('%Y-%m-%d')
        if isinstance(v, (int, float)):
            return (datetime(1899, 12, 30) + timedelta(days=int(v))).strftime('%Y-%m-%d')
        return str(v).strip()[:10]

    def parse_time(v):
        if not v:
            return None
        if isinstance(v, (int, float)):
            total_min = round(v * 24 * 60)
            h, m = divmod(total_min, 60)
            return f"{h:02d}:{m:02d}"
        s = str(v).strip()
        if ':' in s:
            return s[:5]
        return None

    # Normalize rows to common format
    groups = OrderedDict()
    errors = []

    for i, r in enumerate(raw_rows, 2):
        try:
            if is_format_b:
                # Format B: OT_ID(0), Fecha(1), Actividad(2), Trabajador(3), H_ini(4), H_fin(5), Pausa(6),
                #           HBruta(7), HNetas(8), Cant(9), Modalidad(10), Metrica(11), CostoHora(12), CostoMO(13),
                #           ..., Observaciones(17), Campo(18), Estado(19), Dia(20), Mes(21), Ano(22), Equipo(23)
                fecha = parse_fecha(r[1])
                act_name = str(r[2]).strip() if r[2] else ''
                trab_name = str(r[3]).strip() if r[3] else ''
                h_ini = parse_time(r[4])
                h_fin = parse_time(r[5])
                pausa = int(r[6]) if r[6] and isinstance(r[6], (int, float)) else 0
                horas_netas = float(r[8]) if r[8] and isinstance(r[8], (int, float)) else None
                costo_hora_xl = float(r[12]) if r[12] and isinstance(r[12], (int, float)) else None
                costo_mo_xl = float(r[13]) if r[13] and isinstance(r[13], (int, float)) else None
                campo = str(r[18]).strip() if len(r) > 18 and r[18] else ''
                estado = str(r[19]).strip() if len(r) > 19 and r[19] else 'Abierta'
                equipo = str(r[23]).strip() if len(r) > 23 and r[23] else ''
                supervisor = 'Percido Valerio'
            else:
                # Format A (plantilla): with or without OT grouper column
                # With OT col: OT(0), Fecha(1), Actividad(2), Trabajador(3), H_ini(4), H_fin(5),
                #               Pausa(6), Campo(7), Supervisor(8), Equipo(9), Estado(10)
                # Without OT col: Fecha(0), Actividad(1), Trabajador(2), H_ini(3), H_fin(4),
                #                  Pausa(5), Campo(6), Supervisor(7), Equipo(8), Estado(9)
                if has_ot_col and not is_format_b:
                    ot_group = str(r[0]).strip() if r[0] else ''
                    offset = 1
                else:
                    ot_group = ''
                    offset = 0
                fecha = parse_fecha(r[offset])
                act_name = str(r[offset+1]).strip() if r[offset+1] else ''
                trab_name = str(r[offset+2]).strip() if r[offset+2] else ''
                h_ini = parse_time(r[offset+3])
                h_fin = parse_time(r[offset+4])
                pausa = int(r[offset+5]) if r[offset+5] and str(r[offset+5]).strip() not in ('', '\u2014') else 0
                horas_netas = None
                costo_hora_xl = None
                costo_mo_xl = None
                campo = str(r[offset+6]).strip() if len(r) > offset+6 and r[offset+6] else ''
                supervisor = str(r[offset+7]).strip() if len(r) > offset+7 and r[offset+7] else ''
                equipo = str(r[offset+8]).strip() if len(r) > offset+8 and r[offset+8] else ''
                estado = str(r[offset+9]).strip() if len(r) > offset+9 and r[offset+9] else 'Abierta'

            # Clean campo
            if campo in ('Vivero', 'General', 'None', 'Area comun', 'area comun'):
                campo = ''

            # Resolve activity
            act_id = acts.get(act_name) or acts_lower.get(act_name.lower().strip())
            if not act_id:
                errors.append(f"Fila {i}: Actividad '{act_name}' no encontrada")
                continue

            # Resolve worker
            trab_id = trabs.get(trab_name) or trabs_lower.get(trab_name.lower().strip())
            if not trab_id and trab_name and trab_name not in ('(Sin trabajador)', 'None', ''):
                errors.append(f"Fila {i}: Trabajador '{trab_name}' no encontrado")
                continue

            # Validate campo
            if campo and campo not in campos_set:
                errors.append(f"Fila {i}: Campo '{campo}' no existe")
                continue

            # Group key: use ot_group if available (allows multiple OTs with same fecha/act/campo)
            if is_format_b:
                ot_group = str(r[0]).strip()  # OT_ID from original Excel
            elif not has_ot_col:
                ot_group = ''  # no grouper — group by fecha/act/campo

            key = (ot_group, fecha, act_id, campo, supervisor, equipo, estado)
            if key not in groups:
                groups[key] = {'h_ini': h_ini, 'workers': []}

            if trab_id:
                # Calculate hours
                if horas_netas is not None:
                    horas = round(horas_netas, 2)
                elif h_ini and h_fin:
                    h1, m1 = map(int, h_ini.split(':'))
                    h2, m2 = map(int, h_fin.split(':'))
                    mins = (h2*60+m2) - (h1*60+m1) - pausa
                    if mins < 0:
                        mins += 24*60
                    horas = round(mins / 60, 2)
                else:
                    horas = 8.0

                # Use Excel costo if available, else calculate from tarifa
                if costo_hora_xl is not None:
                    costo_hora = round(costo_hora_xl, 2)
                else:
                    act_obj = db.query(models.Actividad).filter(models.Actividad.id_act == act_id).first()
                    tarifa = act_obj.tarifa_jornada if act_obj and act_obj.tarifa_jornada else 530
                    costo_hora = round(tarifa / 8, 2)

                costo_mo = round(costo_mo_xl, 2) if costo_mo_xl is not None else round(costo_hora * horas, 2)

                groups[key]['workers'].append({
                    'trabajador_id': trab_id,
                    'modalidad': 'Jornada',
                    'horas_netas': horas,
                    'hora_inicio': h_ini,
                    'hora_fin': h_fin,
                    'costo_hora': costo_hora,
                    'costo_mo': costo_mo,
                })
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")

    # Create OTs
    created = 0
    for (ot_group, fecha, act_id, campo, supervisor, equipo, estado), data in groups.items():
        try:
            ot_id = get_next("OT", db)
            costo_mo_total = sum(w['costo_mo'] for w in data['workers'])

            orden = models.OrdenTrabajo(
                ot_id=ot_id,
                fecha_ejecucion=fecha,
                hora_inicio=data['h_ini'],
                campo_id=campo or None,
                actividad_id=act_id,
                supervisor=supervisor,
                estado=estado or 'Abierta',
                equipo=equipo,
                costo_mano_obra=costo_mo_total,
                costo_total=costo_mo_total,
            )
            db.add(orden)
            db.flush()

            for w in data['workers']:
                mo = models.OTManoObra(
                    ot_id=ot_id,
                    trabajador_id=w['trabajador_id'],
                    fecha=fecha,
                    hora_inicio=w['hora_inicio'],
                    hora_fin=w['hora_fin'],
                    horas_netas=w['horas_netas'],
                    modalidad='Jornada',
                    costo_hora=w['costo_hora'],
                    costo_mo=w['costo_mo'],
                )
                db.add(mo)

            created += 1
        except Exception as e:
            errors.append(f"OT {fecha}/{act_id}/{campo}: {str(e)}")

    db.commit()
    return {
        "format": "OT_ManoObra" if is_format_b else "Plantilla",
        "created": created,
        "total_rows": len(raw_rows),
        "total_ots": len(groups),
        "errors": errors[:50],
        "error_count": len(errors),
    }
